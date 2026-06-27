import os
import io
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from database import get_db
from models import Project, Template, ProjectTemplate, ProjectVariable, GeneratedFile
from schemas import (
    ProjectCreate, ProjectOut, ProjectDetailOut,
    BindTemplatesRequest, BindTemplatesResponse, ProjectVariableOut,
    VariablesSubmit
)
from services.variable_service import merge_variable_keys
from services.generate_service import clear_project_outputs
import openpyxl

router = APIRouter(prefix="/api/projects", tags=["projects"])


# ── 项目列表 ──────────────────────────────────────────────

@router.get("", response_model=list[ProjectOut])
def list_projects(db: Session = Depends(get_db)):
    return db.query(Project).order_by(Project.created_at.desc()).all()


# ── 创建项目 ──────────────────────────────────────────────

@router.post("", response_model=ProjectOut)
def create_project(body: ProjectCreate, db: Session = Depends(get_db)):
    if not body.name.strip():
        raise HTTPException(status_code=400, detail="项目名称不能为空")
    proj = Project(name=body.name.strip())
    db.add(proj)
    db.commit()
    db.refresh(proj)
    return proj


# ── 项目详情 ──────────────────────────────────────────────

@router.get("/{project_id}", response_model=ProjectDetailOut)
def get_project(project_id: str, db: Session = Depends(get_db)):
    proj = db.get(Project, project_id)
    if not proj:
        raise HTTPException(status_code=404, detail="项目不存在")

    templates = [pt.template for pt in proj.project_templates]

    return ProjectDetailOut(
        id=proj.id,
        name=proj.name,
        status=proj.status,
        created_at=proj.created_at,
        templates=templates,
        variables=proj.variables,
        generated_files=proj.generated_files,
    )


# ── 绑定模板 ──────────────────────────────────────────────

@router.post("/{project_id}/bind", response_model=BindTemplatesResponse)
def bind_templates(
    project_id: str,
    body: BindTemplatesRequest,
    db: Session = Depends(get_db)
):
    proj = db.get(Project, project_id)
    if not proj:
        raise HTTPException(status_code=404, detail="项目不存在")
    if not body.template_ids:
        raise HTTPException(status_code=400, detail="至少选择一个模板")

    # 验证模板是否都存在
    templates = []
    for tid in body.template_ids:
        tmpl = db.get(Template, tid)
        if not tmpl:
            raise HTTPException(status_code=404, detail=f"模板 {tid} 不存在")
        templates.append(tmpl)

    # 清空旧绑定、变量、生成文件记录
    db.query(ProjectTemplate).filter(ProjectTemplate.project_id == project_id).delete()
    db.query(ProjectVariable).filter(ProjectVariable.project_id == project_id).delete()
    db.query(GeneratedFile).filter(GeneratedFile.project_id == project_id).delete()
    db.commit()

    # 删除磁盘上的生成文件
    clear_project_outputs(project_id)

    # 写入新的绑定关系
    for tmpl in templates:
        db.add(ProjectTemplate(project_id=project_id, template_id=tmpl.id))

    # 对所有模板的 variable_keys 做并集去重，写入 ProjectVariable
    merged = merge_variable_keys(templates)
    pv_list = []
    for var_def in merged:
        pv = ProjectVariable(
            project_id=project_id,
            key=var_def["key"],
            label="",
            value="",
            type=var_def.get("type", "text"),
        )
        db.add(pv)
        pv_list.append(pv)

    proj.status = "filling"
    db.commit()

    # 刷新获取 id
    for pv in pv_list:
        db.refresh(pv)

    return BindTemplatesResponse(
        project_id=project_id,
        variables=[ProjectVariableOut.model_validate(pv) for pv in pv_list],
        excel_download_url=f"/api/projects/{project_id}/variables/excel",
    )


# ── 下载变量 Excel 模板 ───────────────────────────────────

@router.get("/{project_id}/variables/excel")
def download_variable_excel(project_id: str, db: Session = Depends(get_db)):
    proj = db.get(Project, project_id)
    if not proj:
        raise HTTPException(status_code=404, detail="项目不存在")

    variables = db.query(ProjectVariable).filter(
        ProjectVariable.project_id == project_id
    ).all()

    if not variables:
        raise HTTPException(status_code=400, detail="该项目尚未绑定模板或无变量")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "变量填写"
    # 表头
    ws.append(["key", "label", "value"])
    # 填充 key 列，label 和 value 留空由用户填写
    for v in variables:
        ws.append([v.key, v.label or "", v.value or ""])

    # 冻结首行
    ws.freeze_panes = "A2"
    # 设置列宽
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 30

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=variables_{project_id}.xlsx"},
    )


# ── 提交变量值（表单 JSON） ───────────────────────────────

@router.post("/{project_id}/variables")
def submit_variables(
    project_id: str,
    body: VariablesSubmit,
    db: Session = Depends(get_db)
):
    proj = db.get(Project, project_id)
    if not proj:
        raise HTTPException(status_code=404, detail="项目不存在")

    # 校验所有变量都填写了 label 和 value
    for item in body.variables:
        if not item.label.strip():
            raise HTTPException(status_code=400, detail=f"变量 {item.key} 的 label 不能为空")
        if not item.value.strip():
            raise HTTPException(status_code=400, detail=f"变量 {item.key} 的 value 不能为空")

    # 更新 DB
    for item in body.variables:
        pv = db.query(ProjectVariable).filter(
            ProjectVariable.project_id == project_id,
            ProjectVariable.key == item.key
        ).first()
        if pv:
            pv.label = item.label.strip()
            pv.value = item.value.strip()

    db.commit()
    return {"message": "变量保存成功", "count": len(body.variables)}


# ── 上传填写好的 Excel 提交变量 ───────────────────────────

@router.post("/{project_id}/variables/upload")
async def upload_variable_excel(
    project_id: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    proj = db.get(Project, project_id)
    if not proj:
        raise HTTPException(status_code=404, detail="项目不存在")

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))  # 跳过表头
    if not rows:
        raise HTTPException(status_code=400, detail="Excel 文件无数据行")

    # 校验并收集数据
    parsed = []
    for i, row in enumerate(rows, start=2):
        key = str(row[0]).strip() if row[0] else ""
        label = str(row[1]).strip() if row[1] else ""
        value = str(row[2]).strip() if row[2] else ""
        if not key:
            continue
        if not label:
            raise HTTPException(status_code=400, detail=f"第 {i} 行 label 不能为空（key={key}）")
        if not value:
            raise HTTPException(status_code=400, detail=f"第 {i} 行 value 不能为空（key={key}）")
        parsed.append({"key": key, "label": label, "value": value})

    if not parsed:
        raise HTTPException(status_code=400, detail="Excel 中没有有效数据")

    # 更新 DB
    for item in parsed:
        pv = db.query(ProjectVariable).filter(
            ProjectVariable.project_id == project_id,
            ProjectVariable.key == item["key"]
        ).first()
        if pv:
            pv.label = item["label"]
            pv.value = item["value"]

    db.commit()
    return {"message": "Excel 导入成功", "count": len(parsed)}
